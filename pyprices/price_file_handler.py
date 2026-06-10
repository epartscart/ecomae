#Универсальный класс для обработки файла прайс-листа (для форматов CSV, TXT, XLS, XLSX)



import file_receiver #Классы для получения файлов (могут использоваться их статические методы)
import os
import price_record #Класс товарной позиции

#Работы с файлами и директориями
from pathlib import Path

import csv
import sys

try:
    csv.field_size_limit(sys.maxsize)
except OverflowError:
    csv.field_size_limit(2**31 - 1)  # Max value for C long on Windows



"""
Для работы с Excel-файлами

Пока выбран именно такой набор библиотек для соответствующих типов экселевских файлов. Они протестированы и работают. В будущем, возможно еще больше оптимизировать работу.
Сейчас xlrd используется для XLS. Такие файлы имеют максимум 65 000 строк и обрабатываются очень быстро.
Возможно, эту же библиотеку использовать и для XLSX, т.к. она быстрее, чем openpyxl. НО, XLSX может содержать очень много строк, например 1М и больше. Поэтому, для чтения XLSX используется openpyxl с режимом read_only, который позволяет работать с большими файлами.
"""
from openpyxl import load_workbook #XLSX
import xlrd #XLS



class PriceFileHandler:
    
    #Для автоматической настройки кодировки и разделителей колонок
    text_delimiters = [';', '\t', ','] #Возможные разделители колонок для текстовых файлов
    text_delimiter_index = 0 #Текущий индекс из массива возможных разделителей
    
    text_files_extensions = ['txt', 'csv'] #Расширения текстовых файлов
    openpyxl_files_extensions = ['xlsx'] #Файлы, обрабатываемые библиотекой openpyxl
    xlrd_files_extensions = ['xls'] #Файлы, обрабатываемые библиотекой xlrd
        
    #excel_files_extensions = ['xls', 'xlsx'] #Расширения экселевских файлов (не используется, т.к. нет хорошей универсальной библиотеки для xls и xlsx)

    # --------------------------------------------------------------
    
    #Конструктор
    def __init__(self, task, filename, db_link, launch_id):
        
        #ID запуска
        self.launch_id = launch_id
        
        #Объект задания
        self.task = task
        
        #Имя одного файла, который сейчас обрабатываем из задания
        self.filename = filename
        
        #Количество импортированных (или обработанных) строк
        self.file_records_handled = 0
        
        #Подключение к БД
        self.db_link = db_link
        
        #Массив для хранения декскрипторов буферных файлов (массив нужен для соответствующих self.import_mode)
        self.buffer_files = []
        
        
        #Режим импорт в БД
        self.import_mode = '1_mysql_inserts'
        """
        1_mysql_inserts - Прямой импорт в БД через INSERT-запросы:
        - без создания буферных файлов
        - без асинхронности (т.е. при работе модуля обновления прайс-листов, обрабатываем все задания и сразу пишем в БД - по исходной блок-схеме)
        
        [X] - пишем буферные файлы. Далее могут быть различные вариации:
        - под каждый исходный файл - отдельный буферный файл (не зависит от количества строк)
        - в одном буфурном файле лимит в N строк (при достижении, начинает создаваться следующий файл)
        - буферные файлы загружаем сразу при работе модуля
        - буферные файлы не загружаем, а отдаем контроллеру их список (например, чтобы затем контроллер мог бы их импортировать асинхронно - запрос-импорт-ответ и так по циклу буферных файлов)
        """
    
    # --------------------------------------------------------------
    
    #Чтобы можно было использовать контекстный менеджер
    def __enter__(self):
        return self
    def __exit__(self, exception_type, exception_value, traceback):
        #Тут можно что-то сделать
        
        #Необходимо закрыть все открытые буферные файлы
        for file in self.buffer_files:
            if not file.closed:
                file.close()
                #print('Успешно закрыли буферный файл')
        
    
    # --------------------------------------------------------------
    
    
    #Метод создания папки tmp/launch_id/buffer_files (куда будут складываться буферные csv перед импортом их в БД)
    def create_buffer_files_folder(self):
        
        if not os.path.isdir( 'tmp/' + str(self.launch_id) + '/buffer_files' ):
            #Создаем директорию
            os.mkdir( 'tmp/' + str(self.launch_id) + '/buffer_files' )
            if not os.path.isdir( 'tmp/' + str(self.launch_id) + '/buffer_files' ):
                #Не создали
                raise Exception('Could not create directory tmp/launch_id('+ str(self.launch_id) +')/buffer_files' )
                return
    
    
    # --------------------------------------------------------------
    
    #Метод старта обработки файла
    def handle_file(self):
        #print("Начали обработку файла " + self.filename + "<br>")
        
        #Инициализация режима импорта
        self.init_import_mode()
        
        #TODO. 20231001. Ниже идет алгоритм чтения файла соответсвующей функцией. Такой, не самый внятный способ определения подходящей функции выработан после многократного тестирования на реальных сайтах. Возможно, требуется проработать еще более широкий алгоритм - сделать механизм try except под кажлый тип файла - и пробовать обрабатывать файл разными способами, пока не обработается (пробовать разные функции, обрабатывать все исключения, смотреть количество строк при отсутствии исплючения и т.д.) - тогда скрипт будет более надежным
        
        
        #Здесь определяем тип файла и вызываем соответствующий метод: handle_text_file() - для текстовых файлов (TXT, CSV); handle_openpyxl_file() - для актуального Excel (XLSX); handle_xlrd_file() - для старого формата Excel (XLS). Каждый из них читает файл соответствующим образом, но, результат одинаков - построчное чтение, создание объектов PriceRecord для каждой строки, передача этих объектов в общий метод импорта (обработки)
        file_extension = file_receiver.file_receiver.get_file_extension(self.filename)
        
        if file_extension in self.text_files_extensions:
            
            #CSV и TXT
            self.handle_text_file()
            
        elif file_extension in self.openpyxl_files_extensions:
            
            #XLSX
            try:
                
                #Сначала пробуем работать через более подходящую для XLSX библиотеку
                self.handle_openpyxl_file()
                
                #Если функция handle_openpyxl_file() не вызвала исключение, то, проверяем количество обработанных строк. Если оно равно 0, то, сами генерируем исключение и ниже попробуем обработать этот файл через другую функцию
                if self.file_records_handled == 0:
                    raise Exception('Failed to read file')
                
            except KeyError as err:
                
                #Если возникло исключение KeyError, то, пробуем обработать этот же файл через универсальную библиотеку. Внимание! Именно под такое исключение ставится библиотека именно версии xlrd==1.2.0
                self.task.other_messages.append('An error occurred while processing via xlsx: ' + str(err) + ', trying via xlrd')
                self.handle_xlrd_file()
                
            except Exception as err:
                
                #Если возникло еще какое-то исключение. Тоже пробуем обработать через туже функцию
                self.task.other_messages.append('An error occurred while processing via xlsx: ' + str(err) + ', trying via xlrd')
                self.handle_xlrd_file()
                
        elif file_extension in self.xlrd_files_extensions:
            
            #XLS
            try:
                
                #Сначала пробуем через xlrd
                self.handle_xlrd_file()
                
            except xlrd.biffh.XLRDError:
                
                #ПРЕДПОЛАГАЕМ, что это на самом деле CSV
                #Если возникло такое исключение - это, как обычно, от кривого поставщика - кривой Excel-файл, не соответствующий структуре xls. Обрабатываем его соответствующим образом
                self.task.other_messages.append("File " + str(self.filename) + " has an incorrect structure. It may not be XLS. Trying to rename it to CSV and trying to process it as CSV.")
                
                
                #Переименовываем файл xls в csv путем добавления расширения в конец имени файла (<filename>.xls.csv). При этом проверяем наличие такого файла (предотвращаем колизии)
                f_name_pref = 1 #Префикс имени файла для исключения колизий на случай, если такой файл уже есть
                #Из цикла выйдем, когда убедимся, что такого файла еще нет
                while Path('tmp/' +str(self.launch_id) + '/' + str(self.task.price_id) + '/' + str(f_name_pref) + self.filename + '.csv').is_file():
                    f_name_pref += 1 #Инкрементируем префикс
                
                #Тут уже само переименование
                os.rename('tmp/' + str(self.launch_id) + '/' + str(self.task.price_id) + '/' + self.filename, 'tmp/' + str(self.launch_id) + '/' + str(self.task.price_id) + '/' + str(f_name_pref) + self.filename+'.csv')
                
                #Переиницализация своей переменной, которую будет смотреть функция self.handle_text_file()
                self.filename = str(f_name_pref) + self.filename + '.csv'
                
                self.handle_text_file()
                
        else:
            raise Exception( 'The file has unsupported type ' + str(file_extension) )
            return
    
    # --------------------------------------------------------------
    
    #Автоопределение кодировки текстового файла
    def detect_encoding(self):
        
        #Исходим из того, что пока на практике встречаются файлы двух возможных кодировках: UTF-8 и ANSI (cp1251)
        
        #Пробуем открыть файл в utf-8
        with open('tmp/' + str(self.launch_id) + '/' + str(self.task.price_id) + '/' + self.filename , 'r', encoding='utf-8') as handled_file:
        
            try:
                #Пробуем прочитать первые 3000 знаков
                filedata = handled_file.read(3000)
                return 'utf-8'
            except:
                return 'cp1251'
        
        return False
    
    # --------------------------------------------------------------
    
    #Метод обработки CSV и TXT
    def handle_text_file(self):
        self.task.other_messages.append("Reading via CSV function")
        
        self.task.other_messages.append("Using auto-detect separator mode")
        
        self.task.cols_delimiter = self.text_delimiters[self.text_delimiter_index]
        
        #Индикация разделителя
        cols_delimiter_to_show = self.task.cols_delimiter
        if cols_delimiter_to_show == '\t':
            cols_delimiter_to_show = 'Tabulation'
        
        self.task.other_messages.append("Separator: " + str(cols_delimiter_to_show) )

        #Механизм автоопределения некорректности файла. Нужен, если будем переименовывать xls/xlsx файл в csv и читать, как csv, либо, если выбран неподходящий разделитель. Если этот файл не является текстовым, то, несколько строк подряд при разделении по cols_delimiter не будут иметь достаточное количество колонок. Считаем, что необходимо 3 колонки или больше.
        less_3_cols_counter = 0 #Счетчик количества строк подряд, в которых менее 3 колонок
        less_3_cols_count_to_detect_case = 20 #Сколько строк подряд должны быть некооректными, чтобы посчитать такой файл некорректным
        
        
        #Определяем кодировку
        file_encoding = False
        #Если пользователь оставил настройку "Авто" - опреляем автоматически
        if self.task.file_encoding == 'auto':
            file_encoding = self.detect_encoding()
            if type(file_encoding) is not str:
                self.task.other_messages.append("It was not possible to detect the encoding automatically. Using the default encoding ANSI. It is recommended to set the encoding explicitly in the price list settings")
                file_encoding = 'cp1251'
        else:
            #Берем явное значение от пользователя
            file_encoding = self.task.file_encoding
        
        
        #Открываем файл прайс-листа на чтение
        with open('tmp/' + str(self.launch_id) + '/' + str(self.task.price_id) + '/' + self.filename , 'r', encoding=file_encoding, errors="ignore") as handled_file:
            self.task.other_messages.append("Opened the price list file for reading" )
            reader = csv.reader(handled_file, delimiter=self.task.cols_delimiter)
            self.task.other_messages.append("Got the file to the CSV library" )
            
            i = 0 #Счетчик прочитанных строк
            #Начинаем построчно читать файл прайс-листа
            for line_list in reader:
                
                #Пропускаем нужное количество строк
                if self.task.cols_to_left > i:
                    i += 1 #Считаем количество прочитанных строк
                    continue
                i += 1 #Считаем количество прочитанных строк

                #Детектор некорректности файла. Если количество обработанных строк уже больше 100, то, считаем, что файл корректный
                if len(line_list) < 3 and i < 100:
                    
                    less_3_cols_counter += 1 #Строка не корректная, считаем
                    
                    #Если достигли предельное количесво некорректных строк
                    if less_3_cols_counter > less_3_cols_count_to_detect_case:
                        
                        #Если еще остались возможные разделители
                        if self.text_delimiter_index < (len(self.text_delimiters)-1):
                            self.text_delimiter_index += 1 #На следующем вызове будем использовать следующий разделитель
                            #ПЕРЕИнициализация режима импорта (чтобы отбросить текущий подготавливаемый запрос к БД)
                            self.init_import_mode()

                            self.task.other_messages.append("The current separator did not fit. Trying another one. Recursive call" )
                            
                            #Рекурсивный вызов
                            self.handle_text_file()
                            return
                        else:
                            raise Exception('This file is not a text file (CSV or TXT) due to the fact that it contains a number of incorrect lines in a row: ' + str(less_3_cols_counter) + '. None of the possible separators worked' )
                    else:
                        #Переходим к следующей строке
                        continue
                else:
                    less_3_cols_counter = 0 #Строка корректна - сбрасываем счетчик
                
                
                #Здесь нужно сформировать словарь из существующих колонок и создать на его основе объект класса PriceRecord
                
                #Создаем пустой словарь для товарной позиции
                price_record_dict = {}
                
                #Цикл по списку имен всех возможных колонок в файле
                for price_field in price_record.PriceRecord.price_fields:
                    #Если в задании указана такая колонка
                    if getattr(self.task, "col_" + price_field) != 0:
                        #Если в файле есть такая колонка физически
                        if getattr(self.task, "col_" + price_field) <= len(line_list):
                            #Указываем в словаре соответствующее значение (БЕРЕМ ПО ИНДЕКСУ (-1) )
                            price_record_dict[ price_field ] = line_list[ getattr(self.task, "col_" + price_field) -1 ]
                
                #Затем объект PriceRecord внутри себя сделает приведение значений всех полей к корректному техническому виду
                priceRecord = price_record.PriceRecord(price_record_dict)
                
                
                #Есть готовый ОБЪЕКТ ТОВАРНОЙ ПОЗИЦИИ из файла. Его уже можно записывать в БД
                self.import_record(priceRecord)
        
        #После прочтения всего файла, опять вызываем метод import_record с None. Чтобы обработчик импорта смог импортировать строки, которые еще не импортированы
        self.import_record(None)
    
    # --------------------------------------------------------------
    """
    #Метод обработки CSV и TXT
    def handle_text_file_OLD(self):
        
        self.task.other_messages.append("Читаем через текстовую функцию")
        
        cols_delimiter_to_show = self.task.cols_delimiter
        if cols_delimiter_to_show == '\t':
            cols_delimiter_to_show = 'Табуляция'
        
        self.task.other_messages.append("Разделитель: " + str(cols_delimiter_to_show) )
        
        
        #Механизм автоопределения некорректности файла. Нужен на случай, если будем переименовывать xls/xlsx файл в csv и читать, как csv. Если этот файл не является текстовым, то, несколько строк подряд при разделении по cols_delimiter не будут иметь достаточное количество колонок. Считаем, что необходимо 3 колонки или больше.
        less_3_cols_counter = 0 #Счетчик количества строк подряд, в которых менее 3 колонок
        less_3_cols_count_to_detect_case = 20 #Сколько строк подряд должны быть некооректными, чтобы посчитать такой файл некорректным
        
        
        
        #Открываем файл прайс-листа на чтение
        with open('tmp/' + str(self.launch_id) + '/' + str(self.task.price_id) + '/' + self.filename , 'r', encoding=self.task.file_encoding) as handled_file:
            
            self.task.other_messages.append("Открыли файл прайс-листа на чтение" )
            
            i = 0 #Счетчик прочитанных строк
            #Начинаем построчно читать файл прайс-листа
            for line in handled_file:
                
                #Пропускаем нужное количество строк
                if self.task.cols_to_left > i:
                    i += 1 #Считаем количество прочитанных строк
                    continue
                
                
                i += 1 #Считаем количество прочитанных строк
                
                #Список (list) на основе строки
                line_list = line.split(self.task.cols_delimiter)
                
                
                
                #Детектор некорректности файла. Если количество обработанных строк уже больше 100, то, считаем, что файл корректный
                if len(line_list) < 3 and i < 100:
                    
                    less_3_cols_counter += 1 #Строка не корректная, считаем
                    
                    #Если достигли предельное количесво некорректных строк
                    if less_3_cols_counter > less_3_cols_count_to_detect_case:
                        raise Exception('Данный файл не является текстовым (CSV или TXT) по причине того, что в нем встретилось подряд некорректных строк в количестве ' + str(less_3_cols_counter) + '. ЛИБО, попробуйте выбрать другой разделитель в настройках прайс-листа' )
                    else:
                        #Переходим к следующей строке
                        continue
                else:
                    less_3_cols_counter = 0 #Строка корректна - сбрасываем счетчик
                
                
                #Здесь нужно сформировать словарь из существующих колонок и создать на его основе объект класса PriceRecord
                
                #Создаем пустой словарь для товарной позиции
                price_record_dict = {}
                
                #Цикл по списку имен всех возможных колонок в файле
                for price_field in price_record.PriceRecord.price_fields:
                    #Если в задании указана такая колонка
                    if getattr(self.task, "col_" + price_field) != 0:
                        #Если в файле есть такая колонка физически
                        if getattr(self.task, "col_" + price_field) <= len(line_list):
                            #Указываем в словаре соответствующее значение (БЕРЕМ ПО ИНДЕКСУ (-1) )
                            price_record_dict[ price_field ] = line_list[ getattr(self.task, "col_" + price_field) -1 ]
                
                #Затем объект PriceRecord внутри себя сделает приведение значений всех полей к корректному техническому виду
                priceRecord = price_record.PriceRecord(price_record_dict)
                
                
                #Есть готовый ОБЪЕКТ ТОВАРНОЙ ПОЗИЦИИ из файла. Его уже можно записывать в БД
                self.import_record(priceRecord)
        
        #После прочтения всего файла, опять вызываем метод import_record с None. Чтобы обработчик импорта смог импортировать строки, которые еще не импортированы
        self.import_record(None)
    """
    # --------------------------------------------------------------
    
    #Метод обработки файла через библиотеку openpyxl (для актуального Excel - XLSX)
    def handle_openpyxl_file(self):
        self.task.other_messages.append('Processing via XLSX function. File named: ' + self.filename)
        #print("Обработка через XLSX функцию<br>")
        
        #Есть особенность перебора ячеек в XLSX файле - цикл по строкам и затем цикл по колонкам в строке. Для оптимизации скорости, сначала готовим массив соответствия имен колонок в классе товарной позиции и номеров колонок из объекта задания
        cols_dict = {} # Словарь в формате col_number (номер колонки в задании):col_name (имя колонки)
        #Цикл по списку имен всех возможных колонок в файле
        for price_field in price_record.PriceRecord.price_fields:
            #Если в задании указана такая колонка
            if getattr(self.task, "col_" + price_field) != 0:
                #Добавляем соответствие номера колонки ее имени
                cols_dict[ int(getattr(self.task, "col_" + price_field)) ] = price_field

        
        #Открываем экселевский файл
        workbook = load_workbook(filename='tmp/' + str(self.launch_id) + '/' + str(self.task.price_id) + '/' + self.filename, read_only=True)
        
        #sheet = workbook.active #Выбрали первый лист (Закомментил. Теперь перебираем все страницы, т.к. их может быть несколько)
        #Цикл по всем страницам файла
        for sheet in workbook:
            
            #Был случай, когда файл читался не с начала. Решение проблемы:
            sheet.reset_dimensions()
            
            #Цикл по строкам страницы
            i = 0 #Счетчик прочитанных строк на странице
            for row in sheet.rows:
                
                #Пропускаем нужное количество строк
                if self.task.cols_to_left > i:
                    i += 1 #Считаем количество прочитанных строк
                    continue
                
                i += 1 #Считаем количество прочитанных строк
                
                #Создаем пустой словарь для товарной позиции
                price_record_dict = {}
                

                #Цикл по колонкам внутри строки
                col_number = 1 #Номер колонки в следующем цикле
                for cell in row:
                    
                    #Если в настройках задания указана колонка с таким номером, указываем ее значение в соответствующий элемент словаря
                    if col_number in cols_dict:
                        price_record_dict[ str(cols_dict[col_number]) ] = str(cell.value)
                    col_number += 1
                
                
                #Затем объект PriceRecord внутри себя сделает приведение значений всех полей к корректному техническому виду
                priceRecord = price_record.PriceRecord(price_record_dict)
                    
                    
                #Есть готовый ОБЪЕКТ ТОВАРНОЙ ПОЗИЦИИ из файла. Его уже можно записывать в БД
                self.import_record(priceRecord)
        
        
        #После прочтения всего файла, опять вызываем метод import_record с None. Чтобы обработчик импорта смог импортировать строки, которые еще не импортированы
        self.import_record(None)
        
        #Закрываем workbook после чтения
        workbook.close()
    
    # --------------------------------------------------------------
    
    def handle_xlrd_file(self):
        #print("Обработка через XLS (старый Excel) функцию<br>")
        
        self.task.other_messages.append('Processing via XLS (old Excel) function. File named: ' + self.filename)
        
        #Есть особенность перебора ячеек в XLS файле - цикл по строкам и затем цикл по колонкам в строке. Для оптимизации скорости, сначала готовим массив соответствия имен колонок в классе товарной позиции и номеров колонок из объекта задания
        cols_dict = {} # Словарь в формате col_number (номер колонки в задании):col_name (имя колонки)
        #Цикл по списку имен всех возможных колонок в файле
        for price_field in price_record.PriceRecord.price_fields:
            #Если в задании указана такая колонка
            if getattr(self.task, "col_" + price_field) != 0:
                #Добавляем соответствие номера колонки ее имени
                cols_dict[ int(getattr(self.task, "col_" + price_field)) ] = price_field
        
        
        #В некоторых услучаях требуется переопределить кодировку. Если возникает ошибка обработки, связанная с кодировкой, то, можем использовать параметр кодировки из настроек прайс-листа для текстовых, чтобы переопределить кодировку этой настройкой.
        file_encoding_override = None #По-умолчанию - не переопределяется
        if not self.task.file_encoding == 'auto':
            file_encoding_override = self.task.file_encoding
        
        
        #self.task.other_messages.append('encoding_override: ' + str(file_encoding_override) )
        
        #Открываем файл
        workbook = xlrd.open_workbook(filename='tmp/' + str(self.launch_id) + '/' + str(self.task.price_id) + '/' + self.filename, encoding_override=file_encoding_override)
        
        #Цикл по листам
        for sheet in workbook.sheets():
        
            #Цикл по строкам на листе (с пропуском нужного количества строк)
            for rx in range( self.task.cols_to_left, sheet.nrows ):
                
                #Создаем пустой словарь для товарной позиции
                price_record_dict = {}
                
                #Цикл по колонкам в строке
                for cx in range(sheet.ncols):
                    #Если в настройках задания указана колонка с таким номером, указываем ее значение в соответствующий элемент словаря
                    if cx+1 in cols_dict:
                        
                        value = str( sheet.cell_value(rowx=rx, colx=cx) )
                        
                        #Проблема, описанная здесь: https://stackoverflow.com/questions/4928629/xlrd-excel-script-converting-n-a-to-42
                        textType = sheet.cell(rx,cx).ctype #Get the type of the cell
                        if textType == 5:
                            value = None
                    
                        price_record_dict[ str(cols_dict[cx+1]) ] = value
                
                
                #Затем объект PriceRecord внутри себя сделает приведение значений всех полей к корректному техническому виду
                priceRecord = price_record.PriceRecord(price_record_dict)
                    
                    
                #Есть готовый ОБЪЕКТ ТОВАРНОЙ ПОЗИЦИИ из файла. Его уже можно записывать в БД
                self.import_record(priceRecord)
        
        
        
        #После прочтения всего файла, опять вызываем метод import_record с None. Чтобы обработчик импорта смог импортировать строки, которые еще не импортированы
        self.import_record(None)
    
    # --------------------------------------------------------------
    
    def init_import_mode(self):
        
        if self.import_mode == '1_mysql_inserts':
            self.records_to_insert = [] #Пустой список кортежей для импорта в БД
            #Подготовленный запрос
            self.sql_query_insert = "INSERT INTO `shop_docpart_prices_data` (`id`, `price_id`, `manufacturer`, `article`, `article_show`, `name`, `exist`, `price`, `time_to_exe`, `storage`, `min_order`, `reviewed`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            
            
        
        #Отладочный режим
        elif self.import_mode == 'debug':
            #Создаем директорию под буферные csv
            self.create_buffer_files_folder()
            
            #Создаем буферный файл (дескриптор указываем в массиве)
            self.buffer_files.append( open('tmp/' + str(self.launch_id) + '/buffer_files/test_buffer.csv', 'w') )
    
    # --------------------------------------------------------------
    
    #Метод импорта текущей строки соответствующим способом. record - это объект PriceRecord. Если record == None, это значит, что файл прочитан полностью, и, далее в зависимости от режима импорта нужно выполнить соответствующие действия (например, для '1_mysql_inserts' импортировать оставшиеся не импортированные позиции)
    def import_record(self, record):
        
        #Режим импортирования через обычные SQL-запросы (INSERT)
        if self.import_mode == '1_mysql_inserts':
            
            #В список кортежей добавляем кортеж для данной товарной позиции
            if record != None:
                self.records_to_insert.append( record.get_like_tuple(self.task.price_id) )
            
            #Если количество кортежей достаточно для запроса, либо в record был получен None (после завершения чтения файла) - выполняем запрос
            if len(self.records_to_insert) >= self.task.rows_per_query or (record == None and len(self.records_to_insert) > 0):
                #SQL-запрос
                cursor = self.db_link.cursor()
                try:
                    cursor.executemany(self.sql_query_insert, self.records_to_insert)
                    
                    #Если дошли до сюда, то, значит SQL-запрос выполнился корректно - добавляем количество импортированных строк
                    self.file_records_handled = self.file_records_handled + len(self.records_to_insert)
                    
                except Exception as err:
                    pass
                    #print( str(err) ) #Здесь можно записать лог (если произошла ошибка выполнения запроса)
                
                #Очищаем список кортежей для данного запроса
                self.records_to_insert.clear()
            
        
        #Отладочный режим
        elif self.import_mode == 'debug':
            if record != None:
                #На основе объекта PriceRecord компонуется новая текстовая строка, которая кладется в буферный файл. Таким образом, чтение исходного файла и формирование буферного будет завершено.
                line_to_buffer_size = record.get_like_table_string(self.task.price_id)
                
                #Нужно определиться, импортировать ли буферные файлы сразу или собирать их в папку tmp/launch_id/buffer_files/, а потом уже в api.py добавить еще один цикл - чисто импорт буферных файлов в MySQL
                
                self.buffer_files[ len(self.buffer_files) - 1 ].write( line_to_buffer_size )
    
    # --------------------------------------------------------------